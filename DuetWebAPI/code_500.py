import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
import scipy as sp
from scipy import stats
from scipy.interpolate import UnivariateSpline

"""
To calibrate the load cell data with MTS data (for 500 N load)
"""

loadcell_data = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/loadcell_data_time_500_1.xlsx")
dataframe_lc = loadcell_data.active
mts_data = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/data_MTS_500_1.xlsx")
dataframe_mts = mts_data.active

loadcell_data_2 = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/loadcell_data_time_500_2.xlsx")
dataframe_lc_2 = loadcell_data_2.active
mts_data_2 = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/data_MTS_500_2.xlsx")
dataframe_mts_2 = mts_data_2.active

loadcell_data_3 = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/loadcell_data_time_500_3.xlsx")
dataframe_lc_3 = loadcell_data_3.active
mts_data_3 = openpyxl.load_workbook(
    "DuetWebAPI/calibration_data/data_MTS_500_3.xlsx")
dataframe_mts_3 = mts_data_3.active


lc = []
t = []
i = 1

# Iterate the loop to read the cell values
for row in range(0, dataframe_lc.max_row):
    for col in dataframe_lc.iter_cols(min_col=1, max_col=1, min_row=None, max_row=None):
        if col[row].value != None:
            lc.append(col[row].value * 4.4482216153)
            i += 1

for row in range(0, dataframe_lc.max_row):
    for col in dataframe_lc.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            t.append(col[row].value / 1000)
            i += 1

lc_mts = []
t_mts = []
i = 1

# Iterate the loop to read the cell values

for row in range(0, dataframe_mts.max_row):
    for col in dataframe_mts.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            lc_mts.append(col[row].value)
            i += 1

for row in range(0, dataframe_mts.max_row):
    for col in dataframe_mts.iter_cols(min_col=3, max_col=3, min_row=None, max_row=None):
        if col[row].value != None:
            t_mts.append(col[row].value)
            i += 1

lc_2 = []
t_2 = []
i = 1

# Iterate the loop to read the cell values
for row in range(0, dataframe_lc_2.max_row):
    for col in dataframe_lc_2.iter_cols(min_col=1, max_col=1, min_row=None, max_row=None):
        if col[row].value != None:
            lc_2.append(col[row].value * 4.4482216153)
            i += 1

for row in range(0, dataframe_lc_2.max_row):
    for col in dataframe_lc_2.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            t_2.append(col[row].value / 1000)
            i += 1

lc_mts_2 = []
t_mts_2 = []
i = 1

# Iterate the loop to read the cell values

for row in range(0, dataframe_mts_2.max_row):
    for col in dataframe_mts_2.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            lc_mts_2.append(col[row].value)
            i += 1

for row in range(0, dataframe_mts_2.max_row):
    for col in dataframe_mts_2.iter_cols(min_col=3, max_col=3, min_row=None, max_row=None):
        if col[row].value != None:
            t_mts_2.append(col[row].value)
            i += 1

lc_3 = []
t_3 = []
i_3 = 1

# Iterate the loop to read the cell values
for row in range(0, dataframe_lc_3.max_row):
    for col in dataframe_lc_3.iter_cols(min_col=1, max_col=1, min_row=None, max_row=None):
        if col[row].value != None:
            lc_3.append(col[row].value * 4.4482216153)
            i += 1

for row in range(0, dataframe_lc_3.max_row):
    for col in dataframe_lc_3.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            t_3.append(col[row].value / 1000)
            i += 1

lc_mts_3 = []
t_mts_3 = []
i = 1

# Iterate the loop to read the cell values

for row in range(0, dataframe_mts_3.max_row):
    for col in dataframe_mts_3.iter_cols(min_col=2, max_col=2, min_row=None, max_row=None):
        if col[row].value != None:
            lc_mts_3.append(col[row].value)
            i += 1

for row in range(0, dataframe_mts_3.max_row):
    for col in dataframe_mts_3.iter_cols(min_col=3, max_col=3, min_row=None, max_row=None):
        if col[row].value != None:
            t_mts_3.append(col[row].value)
            i += 1

f = sp.interpolate.interp1d(t, lc)
t_mts_new = []
for i in range(1, len(t_mts)):
    if t_mts[i] >= t[1]:
        t_mts_new.append(t_mts[i])

lc_new = f(t_mts_new)
f_mts = sp.interpolate.interp1d(t_mts, lc_mts)
lc_mts_new = f_mts(t_mts_new)


f2 = sp.interpolate.interp1d(t_2, lc_2)
t_mts_new_2 = []
for i in range(1, len(t_mts_2)):
    if t_mts_2[i] >= t_2[1]:
        t_mts_new_2.append(t_mts_2[i])

lc_new_2 = f(t_mts_new_2)
f_mts_2 = sp.interpolate.interp1d(t_mts_2, lc_mts_2)
lc_mts_new_2 = f_mts_2(t_mts_new_2)


f3 = sp.interpolate.interp1d(t_3, lc_3)
t_mts_new_3 = []
for i in range(1, len(t_mts_3)):
    if t_mts_3[i] >= t_3[1]:
        t_mts_new_3.append(t_mts_3[i])

lc_new_3 = f(t_mts_new_3)
f_mts_3 = sp.interpolate.interp1d(t_mts_3, lc_mts_3)
lc_mts_new_3 = f_mts_3(t_mts_new_3)

error = lc_new - lc_mts_new
error_2 = lc_new_2 - lc_mts_new_2
error_3 = lc_new_3 - lc_mts_new_3

func_error_lcLoad = stats.linregress(lc_new, error)
func_error_lcLoad_2 = stats.linregress(lc_new_2, error_2)
func_error_lcLoad_3 = stats.linregress(lc_new_3, error_3)


def slope_reader():
    return [func_error_lcLoad.slope,
            func_error_lcLoad_2.slope, func_error_lcLoad_3.slope]


def intercept_reader():
    return [func_error_lcLoad.intercept,
            func_error_lcLoad_2.intercept, func_error_lcLoad_3.intercept]
